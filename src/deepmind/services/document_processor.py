"""
Document processor — extracts text from various file formats.
Supports: PDF, DOCX, TXT, Markdown, Python, JS/TS, YAML, JSON, HTML, XLSX.
"""
import io
import json
from pathlib import Path
from typing import Optional

import structlog

log = structlog.get_logger()


class DocumentProcessor:
    """Extract text content from various file formats."""
    
    SUPPORTED_EXTENSIONS = {
        ".txt", ".md", ".py", ".js", ".ts", ".tsx", ".jsx",
        ".java", ".go", ".rs", ".rb", ".php", ".c", ".cpp", ".h",
        ".yaml", ".yml", ".json", ".toml", ".ini", ".cfg", ".conf",
        ".html", ".htm", ".xml", ".csv", ".sql", ".sh", ".bash",
        ".dockerfile", ".env", ".gitignore", ".editorconfig",
        ".pdf", ".docx", ".xlsx",
    }
    
    def extract_text(self, content: bytes, filename: str, mime_type: str = "") -> str:
        """Extract text from file content."""
        ext = Path(filename).suffix.lower()
        
        try:
            if ext == ".pdf":
                return self._extract_pdf(content)
            elif ext == ".docx":
                return self._extract_docx(content)
            elif ext == ".xlsx":
                return self._extract_xlsx(content)
            elif ext in (".html", ".htm"):
                return self._extract_html(content)
            else:
                # Plain text / code files
                return content.decode("utf-8", errors="replace")
        except Exception as e:
            log.warning("document_extract_error", filename=filename, error=str(e))
            return content.decode("utf-8", errors="replace")
    
    def _extract_pdf(self, content: bytes) -> str:
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(content))
            pages = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages)
        except ImportError:
            log.warning("pypdf not installed, falling back to raw decode")
            return content.decode("utf-8", errors="replace")
    
    def _extract_docx(self, content: bytes) -> str:
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs)
        except ImportError:
            log.warning("python-docx not installed")
            return ""
    
    def _extract_xlsx(self, content: bytes) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            text_parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text_parts.append(f"## Sheet: {sheet}")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        text_parts.append(" | ".join(cells))
            return "\n".join(text_parts)
        except ImportError:
            log.warning("openpyxl not installed")
            return ""
    
    def _extract_html(self, content: bytes) -> str:
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(content, "html.parser")
            for tag in soup(["script", "style", "nav", "footer", "header"]):
                tag.decompose()
            return soup.get_text(separator="\n", strip=True)
        except ImportError:
            return content.decode("utf-8", errors="replace")
    
    def is_supported(self, filename: str) -> bool:
        return Path(filename).suffix.lower() in self.SUPPORTED_EXTENSIONS


_processor: Optional[DocumentProcessor] = None


def get_document_processor() -> DocumentProcessor:
    global _processor
    if _processor is None:
        _processor = DocumentProcessor()
    return _processor
